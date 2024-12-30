import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import threading
import queue
import re
from datetime import datetime

# For Excel reading
try:
    import xlrd  # For old .xls
except ImportError:
    xlrd = None

try:
    import openpyxl  # For .xlsx
    from openpyxl import Workbook
except ImportError:
    openpyxl = None
    Workbook = None

# -----------------------------------------------------------------------
# GLOBALS
# -----------------------------------------------------------------------
input_files = ["file1.xls", "file2.xls", "file3.txt", "file4.xlsx", "file5.txt"]
result_queue = queue.Queue()
default_max_results = 100
searching = False
loaded_keywords = []

root = tk.Tk()
root.title("Stronger Bulk Search (Excel & Text) - Large File Friendly")
root.geometry("1400x900")

# -----------------------------------------------------------------------
# 1. Upload Keywords
# -----------------------------------------------------------------------
def upload_keywords():
    global loaded_keywords

    file_path = filedialog.askopenfilename(
        title="Select Keywords File",
        filetypes=[("All Files", "*.*")]
    )
    if not file_path:
        messagebox.showwarning("No File Selected", "No file was selected.")
        return

    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as infile:
            lines = [line.strip() for line in infile if line.strip()]

        if not lines:
            messagebox.showinfo("Empty File", "The selected file has no valid keywords.")
            return

        loaded_keywords = lines
        messagebox.showinfo(
            "Keywords Loaded",
            f"Loaded {len(lines)} keyword(s) from:\n{file_path}"
        )

        # Display them
        keywords_text.delete("1.0", tk.END)
        keywords_text.insert(tk.END, "Loaded Keywords:\n\n")
        for kw in loaded_keywords:
            keywords_text.insert(tk.END, kw + "\n")

    except Exception as e:
        messagebox.showerror("File Error", f"Could not read file:\n{e}")

# -----------------------------------------------------------------------
# 2. Generator: Parse file line-by-line (or row-by-row) without storing all
# -----------------------------------------------------------------------
def parse_file_generator(file_path):
    """
    Yields each line/row from the file_path without storing everything in memory.
    For Excel, we convert each row to a tab-joined string, e.g. "192.168.1.1\ta\tdoha".

    This approach is more memory-efficient for huge files.
    """
    extension = file_path.lower().split(".")[-1]

    # 2A. XLS old Excel
    if extension == "xls" and xlrd is not None:
        try:
            workbook = xlrd.open_workbook(file_path)
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                for row_idx in range(sheet.nrows):
                    row_values = sheet.row_values(row_idx)
                    # Convert to one string with tabs
                    yield "\t".join(str(v) for v in row_values)
        except Exception as e:
            # Return an error line so we can handle it gracefully
            yield f"ERROR_PARSING_XLS: {e}"
        return

    # 2B. XLSX new Excel
    elif (extension == "xlsx" or extension == "xlsm") and openpyxl is not None:
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    # row is a tuple of cell values. Filter out None, convert to str, join with tabs.
                    row_str = "\t".join(str(v) for v in row if v is not None)
                    yield row_str
        except Exception as e:
            yield f"ERROR_PARSING_XLSX: {e}"
        return

    # 2C. Otherwise, treat as plain text
    else:
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as infile:
                for raw_line in infile:
                    # Strip newlines, yield directly
                    yield raw_line.rstrip("\n\r")
        except Exception as e:
            yield f"ERROR_PARSING_TXT: {e}"

# -----------------------------------------------------------------------
# 3. Search a Single File for One Keyword
# -----------------------------------------------------------------------
def search_file_for_keyword(file_name, keyword, max_results):
    """
    Reads the file using parse_file_generator (line by line),
    returns a list of (line_number, file_name, line_content) 
    for lines containing `keyword`. Done in a streaming way.
    """
    matches = []
    line_number = 0

    # Regex for case-insensitive search
    pattern = re.compile(re.escape(keyword), re.IGNORECASE)

    for line in parse_file_generator(file_name):
        line_number += 1
        # If there's an error line from parse_file_generator, handle it
        if line.startswith("ERROR_PARSING_"):
            # Return a single entry indicating the file error
            return [(None, file_name, line)]
        # If we reached max results, stop
        if len(matches) >= max_results:
            break
        # Check if this line contains the keyword
        if pattern.search(line):
            matches.append((line_number, file_name, line.strip()))
    return matches

# -----------------------------------------------------------------------
# 4. Main Search: Loop Over Each Keyword (still per-keyword approach)
# -----------------------------------------------------------------------
def refine_lines():
    global searching
    if searching:
        return

    if not loaded_keywords:
        tree.delete(*tree.get_children())
        messagebox.showwarning("No Keywords", "No keywords loaded. Please upload a keywords file first.")
        return

    selected_files = []
    for file_var, file_name in zip(file_vars, input_files):
        if file_var.get() and os.path.exists(file_name):
            selected_files.append(file_name)

    if not selected_files:
        messagebox.showwarning("No Files", "No valid files selected or files do not exist.")
        return

    max_results = default_max_results
    if result_limit_entry.get().isdigit():
        max_results = int(result_limit_entry.get())

    progress_var.set(0)
    progress_bar["maximum"] = 100
    progress_bar["value"] = 0

    def process_search():
        global searching
        searching = True

        final_results = []
        total_steps = len(loaded_keywords)
        step_value = 100 / total_steps if total_steps else 100

        for keyword in loaded_keywords:
            keyword_found_any = False

            for file_name in selected_files:
                matches = search_file_for_keyword(file_name, keyword, max_results)
                if matches:
                    # If there's at least one match or error line, we consider this "found"
                    # except for "Error" lines. We'll add them anyway
                    keyword_found_any = True
                    final_results.extend(matches)

            if not keyword_found_any:
                # No match across any file -> "keyword\tNo data found"
                content_str = f"{keyword}\tNo data found"
                final_results.append((None, "(NoFile)", content_str))

            # Update progress after each keyword
            progress_var.set(progress_var.get() + step_value)
            progress_bar["value"] = progress_var.get()
            root.update_idletasks()

        result_queue.put(final_results)
        searching = False

    threading.Thread(target=process_search, daemon=True).start()
    root.after(100, update_results)

# -----------------------------------------------------------------------
# 5. Update Results Table
# -----------------------------------------------------------------------
def update_results():
    try:
        results = result_queue.get_nowait()
        tree.delete(*tree.get_children())

        file_summary = {}
        if results:
            for idx, (line_number, file_name, line_content) in enumerate(results, start=1):
                line_str = line_number if line_number else "N/A"
                tree.insert("", "end", values=(idx, line_str, file_name, line_content))

                # If it's a real match (line_number != None and not error lines)
                if line_number is not None and not line_content.startswith("ERROR_PARSING_"):
                    file_summary[file_name] = file_summary.get(file_name, 0) + 1
        else:
            tree.insert("", "end", values=("No data found!!", "", "", "No results at all."))

        if file_summary:
            summary_text = "Total Refined Lines: " + ", ".join(f"{f}: {c}" for f, c in file_summary.items())
        else:
            summary_text = "Total Refined Lines: 0"

        total_label.config(text=summary_text)
        progress_var.set(100)
        progress_bar["value"] = 100

    except queue.Empty:
        if searching:
            root.after(100, update_results)

# -----------------------------------------------------------------------
# 6. Clear Results
# -----------------------------------------------------------------------
def clear_results():
    tree.delete(*tree.get_children())
    total_label.config(text="Total Refined Lines: 0")
    progress_var.set(0)

# -----------------------------------------------------------------------
# 7. Save Results (Split Tabs for Excel Columns)
# -----------------------------------------------------------------------
def save_results():
    displayed_results = tree.get_children()
    if not displayed_results:
        messagebox.showinfo("Save Error", "No results to save.")
        return

    if Workbook is None:
        messagebox.showerror(
            "Missing openpyxl",
            "OpenPyXL not installed or import failed. Cannot save as Excel."
        )
        return

    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"scan_results_{now_str}.xlsx"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"

        header = ["No.", "Line", "File", "Content(Cell1)", "Content(Cell2)", "Content(Cell3)", "..."]
        ws.append(header)

        for item in displayed_results:
            row_values = tree.item(item, "values")  # (No., Line, File, Content)
            no_ = row_values[0]
            line_ = row_values[1]
            file_ = row_values[2]
            content_ = row_values[3]

            if "\t" in content_:
                split_cells = content_.split("\t")
            else:
                split_cells = [content_]

            combined = [no_, line_, file_] + split_cells
            ws.append(combined)

        wb.save(output_file)
        messagebox.showinfo("Saved", f"Results saved as Excel file:\n{output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving: {e}")

# -----------------------------------------------------------------------
# 8. GUI Setup
# -----------------------------------------------------------------------
intro_label = tk.Label(
    root,
text=(
    "Bulk Search (Large File Support):\n"
    "1) Upload a .txt keyword file.\n"
    "2) Click 'Search' to stream line-by-line.\n"
    "3) 'Save Results' to get result on xls file.\n"
"# Unmatched keywords appear as 'No data found' #.\n"
),

    font=("Aptos", 12, "bold"),
    fg="steel blue",
    justify="left"
)
intro_label.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky="w")

upload_button = tk.Button(root, text="Upload Keywords File", command=upload_keywords, bg="lightblue", width=32, font=("Arial", 12, "bold"))
upload_button.grid(row=1, column=0, padx=10, pady=5, sticky="w")

result_limit_label = tk.Label(root, text="Max Results (default 100):", font=("Arial", 12))
result_limit_label.grid(row=1, column=1, padx=5, pady=5, sticky="e")
result_limit_entry = tk.Entry(root, width=10, font=("Arial", 12))
result_limit_entry.grid(row=1, column=2, padx=5, pady=5, sticky="w")
result_limit_entry.insert(0, str(default_max_results))

keywords_text = tk.Text(root, wrap="word", width=50, height=7, font=("Arial", 12))
keywords_text.grid(row=2, column=0, padx=10, pady=5, sticky="nw", rowspan=2)
keywords_text.insert("1.0", "(Uploaded keywords will appear here)")

file_frame = tk.Frame(root)
file_frame.grid(row=2, column=1, columnspan=2, padx=10, pady=5, sticky="nw")

file_vars = [tk.IntVar(value=1) for _ in input_files]
for idx, file_name in enumerate(input_files):
    color = "green" if os.path.exists(file_name) else "red"
    chk = tk.Checkbutton(file_frame, text=file_name, variable=file_vars[idx], font=("Arial", 10), fg=color)
    chk.pack(anchor="w", padx=5)

button_frame = tk.Frame(root)
button_frame.grid(row=4, column=0, columnspan=3, pady=10)

search_button = tk.Button(button_frame, text="Search", command=refine_lines, bg="lightgreen", width=15, font=("Arial", 12, "bold"))
search_button.grid(row=0, column=0, padx=5)

save_button = tk.Button(button_frame, text="Save", command=save_results, bg="lightyellow", width=15, font=("Arial", 12, "bold"))
save_button.grid(row=0, column=1, padx=5)

clear_button = tk.Button(button_frame, text="Clear", command=clear_results, bg="lightcoral", width=15, font=("Arial", 12, "bold"))
clear_button.grid(row=0, column=2, padx=5)

tree_frame = tk.Frame(root)
tree_frame.grid(row=5, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

tree = ttk.Treeview(tree_frame, columns=("No.", "Line", "File", "Content"), show="headings")
tree.heading("No.", text="No.")
tree.heading("Line", text="Line")
tree.heading("File", text="File")
tree.heading("Content", text="Content")

tree.column("No.", width=60, anchor="center")
tree.column("Line", width=100, anchor="center")
tree.column("File", width=200, anchor="center")
tree.column("Content", width=900, anchor="w")

style = ttk.Style()
style.configure("Treeview", font=("Arial", 14))
style.configure("Treeview.Heading", font=("Arial", 14, "bold"))

scrollbar_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
tree.configure(yscrollcommand=scrollbar_y.set)

scrollbar_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
tree.configure(xscrollcommand=scrollbar_x.set)

tree.pack(fill=tk.BOTH, expand=True)

progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100, length=400)
progress_bar.grid(row=6, column=0, columnspan=3, pady=10)

total_label = tk.Label(root, text="Total Refined Lines: 0", font=("Arial", 12), fg="blue")
total_label.grid(row=7, column=0, columnspan=3, pady=5)

root.grid_rowconfigure(5, weight=1)
root.grid_columnconfigure(2, weight=1)
root.mainloop()
